#!/usr/bin/env python3
"""
Enrichit chaque fichier XLSX gold (dans outputs/) en ajoutant 7 colonnes binaires
(ironie, insulte, emoji, mépris / haine, argot, abréviation, interjection)
basées sur la détection de mots-clés dans les justifications des fichiers JSONL
associés (claude + gemini).

Utilise openpyxl (pas pandas) pour rester léger et rapide.
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook

# ── Configuration ─────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
OUTPUTS_DIR = BASE_DIR / "outputs"

# Regex pour extraire le texte TARGET depuis le prompt JSONL
# (identique à generate_marqueurs_html.py)
TARGET_RE = re.compile(
    r'TARGET:\s*\[.*?\]\s*\(role=[^)]*\)\s*\(time=[^)]*\)\s*"(.+)"\s*$',
    re.MULTILINE,
)

# Colonnes à ajouter et leurs mots-clés associés (recherche case-insensitive)
KEYWORD_COLUMNS = {
    "ironie":          ["ironie", "ironique", "sarcastique"],
    "insulte":         ["insulte", "insultes"],
    "emoji":           ["émoticône", "emoji"],
    "mépris / haine":  ["mépris", "méprisant", "haine", "haineux", "dégoût", "animosité"],
    "argot":           ["argot", "argotique", "familier", "vulgaire"],
    "abréviation":     ["abréviation"],
    "interjection":    ["interjection"],
}

NEW_COLUMN_NAMES = list(KEYWORD_COLUMNS.keys())


# ── Extraction des justifications depuis les JSONL ────────────────────────────

def extract_target_text(prompt: str) -> str:
    """Extrait le texte TARGET depuis le champ prompt d'une ligne JSONL."""
    m = TARGET_RE.search(prompt)
    return m.group(1).strip() if m else ""


def load_justifications_for_folder(folder: Path) -> dict[str, str]:
    """
    Charge tous les fichiers JSONL d'un dossier et construit un mapping :
        texte_normalisé -> concaténation de toutes les justifications
    """
    text_to_justifs: dict[str, list[str]] = {}

    for jsonl_file in sorted(folder.glob("*.jsonl")):
        print(f"    Lecture {jsonl_file.name} ...")
        with open(jsonl_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    row = json.loads(line)
                except json.JSONDecodeError:
                    continue

                # Extraire le texte cible
                prompt = row.get("prompt", "")
                target_text = extract_target_text(prompt)
                if not target_text:
                    continue

                key = target_text.strip()

                # Inscrire la clé même si elle n'a pas de justifications (pour compter les matches)
                if key not in text_to_justifs:
                    text_to_justifs[key] = []

                # Extraire les justifications depuis parsed_json.sitemo_units
                parsed = row.get("parsed_json")
                if not parsed:
                    continue
                units = parsed.get("sitemo_units", [])
                for unit in units:
                    justif = unit.get("justification", "")
                    if justif:
                        text_to_justifs.setdefault(key, []).append(justif)

    # Concaténer toutes les justifications par texte en un seul blob
    return {k: " ".join(v) for k, v in text_to_justifs.items()}


def detect_keywords(justification_blob: str) -> dict[str, int]:
    """
    Pour un blob de justifications concaténées, retourne un dict
    colonne -> 0/1 selon la présence des mots-clés (case-insensitive).
    """
    blob_lower = justification_blob.lower()
    result = {}
    for col_name, keywords in KEYWORD_COLUMNS.items():
        found = any(kw.lower() in blob_lower for kw in keywords)
        result[col_name] = 1 if found else 0
    return result


# ── Mise à jour des fichiers XLSX ─────────────────────────────────────────────

def process_folder(folder: Path):
    """Traite un sous-dossier : charge les JSONL, met à jour le XLSX gold."""
    # Trouver le fichier XLSX gold
    xlsx_files = list(folder.glob("*_gold_flat.xlsx"))
    if not xlsx_files:
        print(f"  ⚠ Pas de fichier gold XLSX dans {folder.name}, ignoré.")
        return
    xlsx_path = xlsx_files[0]

    print(f"  📂 Dossier : {folder.name}")
    print(f"    Fichier gold : {xlsx_path.name}")

    # 1) Charger les justifications depuis les JSONL
    text_to_justifs = load_justifications_for_folder(folder)
    print(f"    → {len(text_to_justifs)} textes avec justifications extraits")

    # 2) Ouvrir le XLSX en lecture-écriture
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Récupérer les en-têtes existants
    headers = [cell.value for cell in ws[1]]
    n_existing_cols = len(headers)

    # Trouver l'index (1-indexed) de la colonne TEXT
    try:
        text_col_idx = headers.index("TEXT") + 1  # openpyxl est 1-indexed
    except ValueError:
        print(f"    ⚠ Colonne 'TEXT' introuvable dans {xlsx_path.name}, ignoré.")
        wb.close()
        return

    # Vérifier si les colonnes existent déjà, les supprimer si oui
    existing_new_cols = [h for h in headers if h in NEW_COLUMN_NAMES]
    if existing_new_cols:
        print(f"    ℹ Colonnes déjà présentes ({existing_new_cols}), elles seront écrasées.")
        # On doit supprimer les anciennes colonnes (de droite à gauche pour garder les indices)
        cols_to_delete = sorted(
            [headers.index(h) + 1 for h in existing_new_cols],
            reverse=True,
        )
        for col_idx in cols_to_delete:
            ws.delete_cols(col_idx)
        # Recalculer les headers après suppression
        headers = [cell.value for cell in ws[1]]
        n_existing_cols = len(headers)
        text_col_idx = headers.index("TEXT") + 1

    # 3) Ajouter les en-têtes des nouvelles colonnes
    for j, col_name in enumerate(NEW_COLUMN_NAMES):
        ws.cell(row=1, column=n_existing_cols + 1 + j, value=col_name)

    # 4) Pour chaque ligne de données, calculer les flags
    n_matched = 0
    n_total = 0
    for row_idx in range(2, ws.max_row + 1):
        text_value = ws.cell(row=row_idx, column=text_col_idx).value
        if text_value is None:
            # Écrire des 0 partout
            for j in range(len(NEW_COLUMN_NAMES)):
                ws.cell(row=row_idx, column=n_existing_cols + 1 + j, value=0)
            continue

        n_total += 1
        text_key = str(text_value).strip()

        # Chercher le texte dans le mapping
        # Utiliser 'None' comme fallback pour distinguer "non trouvé" de "trouvé mais vide"
        justif_blob = text_to_justifs.get(text_key, None)

        if justif_blob is not None:
            n_matched += 1
            # Détecter les mots-clés
            flags = detect_keywords(justif_blob)
        else:
            flags = {col_name: 0 for col_name in NEW_COLUMN_NAMES}

        # Écrire les valeurs
        for j, col_name in enumerate(NEW_COLUMN_NAMES):
            ws.cell(row=row_idx, column=n_existing_cols + 1 + j, value=flags[col_name])

    # 5) Sauvegarder (écraser)
    wb.save(xlsx_path)
    wb.close()

    print(f"    → {n_matched}/{n_total} textes matchés avec les JSONL")
    print(f"    ✅ Fichier mis à jour : {xlsx_path.name}")

    # 6) Vérification rapide : compter les 1 par colonne
    wb2 = load_workbook(xlsx_path, read_only=True)
    ws2 = wb2.active
    headers2 = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
    col_indices = {
        name: headers2.index(name)
        for name in NEW_COLUMN_NAMES
        if name in headers2
    }
    counts = {name: 0 for name in NEW_COLUMN_NAMES}
    data_rows = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        data_rows += 1
        for name, idx in col_indices.items():
            if row[idx] == 1:
                counts[name] += 1
    wb2.close()

    print(f"    📊 Résumé ({data_rows} lignes) :")
    for name in NEW_COLUMN_NAMES:
        print(f"       {name:20s} : {counts[name]:4d} / {data_rows}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Enrichissement des fichiers gold XLSX avec colonnes binaires")
    print("=" * 60)

    if not OUTPUTS_DIR.is_dir():
        print(f"❌ Dossier outputs/ introuvable : {OUTPUTS_DIR}")
        return

    folders = sorted([d for d in OUTPUTS_DIR.iterdir() if d.is_dir()])
    print(f"Dossiers trouvés : {[f.name for f in folders]}\n")

    for folder in folders:
        process_folder(folder)
        print()

    print("🎉 Terminé !")


if __name__ == "__main__":
    main()
